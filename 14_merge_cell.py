from openpyxl import load_workbook

wb = load_workbook("new.xlsx")
ws = wb.active

ws.unmerge_cells()
ws["B2"].value = "MERGED"

wb.save("new.xlsx")
wb.close()