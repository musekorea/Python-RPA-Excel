from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

ws["A1"]=10
ws["A2"]=20
ws["A3"]=30

ws["A6"] = "=SUM(A1:A3)"

wb.save("new.xlsx")
wb.close()