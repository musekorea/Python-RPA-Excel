from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()
ws.title="My Sheet"
ws.sheet_properties.tabColor="ff0000"

ws1 = wb.create_sheet("New Sheet1",1)
ws2 = wb.create_sheet("New Sheet2")

ws1["A1"]="TEST"
copied_sheet = wb.copy_worksheet(ws1)
copied_sheet.title="copied"

print(wb.sheetnames)

wb.save("test.xlsx")
wb.close()