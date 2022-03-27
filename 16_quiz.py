from re import L
from tracemalloc import start
from openpyxl import Workbook
import Score_data

wb = Workbook()
ws = wb.active

ws.append(Score_data.title)
for score in Score_data.scores:
  ws.append(score)

ws["H1"] = "총점"
ws["I1"] = "성적"

for index, cell in enumerate(ws["D"]):
  if index==0:
    continue
  else: 
    cell.value=10
  
for index, score in enumerate( Score_data.scores, start=2):
  ws.cell(row=index, column=8, value=f"=SUM(B{index}:G{index})")
  sum_val = sum(score[1:]) - score[3] + 10
  if sum_val>=90:
    grade="A"
  elif sum_val>=80:
    grade="B"
  elif sum_val>=70:
    grade="C"
  else: 
    grade="D"
  ws.cell(row=index, column=9).value = grade
  if score[1]<5:
    ws.cell(row=index, column=9).value = "F"
  



wb.save("score.xlsx")
wb.close()