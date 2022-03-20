from json import load
from openpyxl import load_workbook

wb = load_workbook("./test.xlsx")
ws = wb.active

ws.move_range("D1:D11", rows=0, cols=-2)
#ws["B1"].value = "English"

wb.save("./test.xlsx")
wb.close()
