from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("./test.xlsx")
ws = wb.active

for row in ws.rows:
  for cell in row: 
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if cell.column==1:
      continue
    if isinstance(cell.value, int) and cell.value>90:
      cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
      cell.font = Font(color="FF0000")

ws.freeze_panes = "B2"

wb.save("new.xlsx")
wb.close()

 