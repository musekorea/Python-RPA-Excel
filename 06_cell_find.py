from openpyxl import load_workbook

wb = load_workbook("./test.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=1):
  if  row[1].value=="English":
    row[1].value = "Korean"
    
wb.save("test.xlsx")
wb.close()