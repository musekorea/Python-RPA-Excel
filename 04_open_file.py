from openpyxl import load_workbook

wb = load_workbook("test.xlsx")
ws = wb.active

for row in range(1,ws.max_row+1):
  for col in range(1,ws.max_column+1):
    print(ws.cell(row=row, column=col).value, end=" ")
  print()

wb.close()